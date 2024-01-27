# es에 데이터 적재
from elasticsearch import Elasticsearch
from elasticsearch import helpers
import pandas as pd
import schedule
import json
import xlrd
from openpyxl import load_workbook

# 엑셀을 json 데이터로 변환
def convert_excel_to_json():
    # excel 파일을 json 데이터로 변환
    file_path = r"C:\dev\hope_project\medical_dataset\medicine_8th.xlsx"
    wb = load_workbook(file_path)
    sheet = wb.active
    print("sheet : " , sheet)
    data_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {
            'code': row[0],
            'name': row[1],
            'name_en': row[2],
            'company': row[3],
            'company_en': row[4],
            'general': row[5],
            'appearance': row[6],
            'ingredient': row[7],
            'ingredient_en': row[8],
            'method': row[9],
            'period': row[10],
            'insurance': row[11],
            'additive': row[12],
            'image': row[13],
            'effect': row[14],
            'usages': row[15],
            'precautions': row[16],
        }
        data_list.append(data)
    json_data = json.dumps(data_list, ensure_ascii=False)
    with open(r"../json_file/medicine.json", 'w+' , encoding='utf-8') as f:
        f.write(json_data)

# es에 데이터 보내는 작업
def send_to_elasticsearch():
    # Elasticsearch 클라이언트 생성
    es = Elasticsearch("http://localhost:9200")

    # 인덱스 생성
    index_name = "medicine"
    file_path = r"../json_file/medicine.json"
    with open(file_path, "r", encoding="utf-8") as file:
        medicine_data = json.load(file)

    # 각 의약품 항목을 별도의 문서로 인덱싱
    for medicine in medicine_data:
        try:
            # 데이터를 전송하기 전에 데이터 확인
            print(f"Indexing data: {medicine}")
            # es.index() : 데이터를 엘라스틱 서치에 저장. 이 작업을 인덱싱이라고 함
            response = es.index(index=index_name, body=medicine)
            print(response)
        except Exception as e:
            print(f"Error indexing document: {e}")
    print("Data send to Elasticsearch")


send_to_elasticsearch()

# 스케줄러 예시
# scheduler.add_job(func=send_to_elasticsearch, trigger="cron", minute='*/1', id="get_movie")
# app.py에 포스트맨 사용을 위한 GET 추가
# app.add_url_rule('/api/elastic', 'send_to_elasticsearch', send_to_elasticsearch, methods=['GET'])
# 포스트 맨에서는 GET 방식으롷 http://127.0.0.1:5000/api/elastic


